"""Importación masiva de estudiantes y grupos desde CSV o Excel."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from django.core.management.color import no_style
from django.db import connection, transaction, IntegrityError
from openpyxl import Workbook, load_workbook
from rest_framework.exceptions import ValidationError

from apps.usuarios.models import Usuario

from .models import Estudiante, Grupo, InscripcionGrupo, Materia

_MODELOS_CON_SECUENCIA = (Estudiante, Materia, Grupo, InscripcionGrupo)


def _sincronizar_secuencias_academico() -> None:
    """Alinea secuencias PostgreSQL tras cargas con IDs explícitos (fixtures, seeds)."""
    if connection.vendor != 'postgresql':
        return
    statements = connection.ops.sequence_reset_sql(no_style(), list(_MODELOS_CON_SECUENCIA))
    with connection.cursor() as cursor:
        for sql in statements:
            cursor.execute(sql)


def _get_or_create_estudiante(
    correo: str,
    defaults: dict,
    docente,
    stats: dict,
) -> tuple[Estudiante, bool]:
    """Obtiene o crea un estudiante por correo; repara secuencia si PostgreSQL falla por PK."""
    correo = correo.lower().strip()
    existente = Estudiante.objects.filter(correo=correo).first()
    if existente:
        stats['estudiantes_vinculados'] += 1
        _vincular_estudiante_docente(existente, docente)
        return existente, False

    try:
        estudiante = Estudiante.objects.create(correo=correo, **defaults)
    except IntegrityError:
        _sincronizar_secuencias_academico()
        existente = Estudiante.objects.filter(correo=correo).first()
        if existente:
            stats['estudiantes_vinculados'] += 1
            _vincular_estudiante_docente(existente, docente)
            return existente, False
        estudiante = Estudiante.objects.create(correo=correo, **defaults)

    stats['estudiantes_creados'] += 1
    _vincular_estudiante_docente(estudiante, docente)
    return estudiante, True

EXTENSIONES = ('.csv', '.xlsx', '.xlsm')
SEPARADOR_ESTUDIANTES = re.compile(r'[;,]')


def _normalizar_header(valor: Any) -> str:
    if valor is None:
        return ''
    return str(valor).strip().lower().replace(' ', '_').replace('í', 'i').replace('ó', 'o')


def _celda_texto(valor: Any) -> str:
    if valor is None:
        return ''
    return str(valor).strip()


def leer_filas_archivo(archivo) -> list[dict]:
    """Lee la primera hoja / CSV y devuelve filas como dict con clave `_fila`."""
    if hasattr(archivo, 'seek'):
        archivo.seek(0)
    nombre = (getattr(archivo, 'name', '') or '').lower()
    content_type = (getattr(archivo, 'content_type', '') or '').lower()
    if nombre.endswith('.csv') or content_type in ('text/csv', 'application/csv'):
        return _leer_csv(archivo)
    if (
        nombre.endswith(('.xlsx', '.xlsm'))
        or 'spreadsheetml' in content_type
        or content_type == 'application/octet-stream' and nombre.endswith('.xlsx')
    ):
        return _leer_xlsx(archivo)
    raise ValidationError({
        'archivo': f'Formato no soportado. Usa: {", ".join(EXTENSIONES)}',
    })


def _leer_csv(archivo) -> list[dict]:
    if hasattr(archivo, 'seek'):
        archivo.seek(0)
    raw = archivo.read()
    if isinstance(raw, bytes):
        texto = raw.decode('utf-8-sig')
    else:
        texto = raw
    reader = csv.DictReader(io.StringIO(texto))
    if not reader.fieldnames:
        raise ValidationError({'archivo': 'El archivo CSV no tiene encabezados.'})
    claves = [_normalizar_header(h) for h in reader.fieldnames]
    filas: list[dict] = []
    for idx, row in enumerate(reader, start=2):
        valores = list(row.values())
        if all(not _celda_texto(v) for v in valores):
            continue
        fila = {'_fila': idx}
        for i, clave in enumerate(claves):
            if clave:
                fila[clave] = valores[i] if i < len(valores) else ''
        filas.append(fila)
    return filas


def _leer_xlsx(archivo) -> list[dict]:
    if hasattr(archivo, 'seek'):
        archivo.seek(0)
    wb = load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active
    iterador = ws.iter_rows(values_only=True)
    encabezados = next(iterador, None)
    if not encabezados:
        raise ValidationError({'archivo': 'El Excel no tiene encabezados.'})
    claves = [_normalizar_header(h) for h in encabezados]
    filas: list[dict] = []
    for idx, row in enumerate(iterador, start=2):
        if all(_celda_texto(c) == '' for c in row):
            continue
        fila = {'_fila': idx}
        for i, clave in enumerate(claves):
            if clave:
                fila[clave] = row[i] if i < len(row) else ''
        filas.append(fila)
    wb.close()
    return filas


def _resolver_materia(nombre: str, docente, stats: dict) -> Materia | None:
    nombre = nombre.strip()
    if not nombre:
        return None
    materia, creada = Materia.objects.get_or_create(
        docente=docente,
        nombre=nombre,
        defaults={'activo': True},
    )
    if creada:
        stats['materias_creadas'] += 1
    return materia


def _partir_nombre(nombre: str) -> tuple[str, str]:
    partes = nombre.split(None, 1)
    if not partes:
        return '', ''
    if len(partes) == 1:
        return partes[0], ''
    return partes[0], partes[1]


def _vincular_estudiante_docente(estudiante: Estudiante, docente) -> None:
    estudiante.docentes.add(docente)


def _inscribir_en_grupo(grupo: Grupo, estudiante: Estudiante, stats: dict) -> None:
    _, creada = InscripcionGrupo.objects.get_or_create(grupo=grupo, estudiante=estudiante)
    if creada:
        stats['inscripciones'] += 1


def _procesar_fila_estudiante(fila: dict, docente, stats: dict) -> None:
    correo = _celda_texto(fila.get('correo')).lower()
    nombre = _celda_texto(fila.get('nombre'))
    if not correo or not nombre:
        raise ValidationError('Las columnas nombre y correo son obligatorias.')

    first_name, last_name = _partir_nombre(nombre)
    identificacion = _celda_texto(fila.get('identificacion'))
    estudiante, creado = _get_or_create_estudiante(
        correo,
        {
            'first_name': first_name,
            'last_name': last_name,
            'identificacion': identificacion,
            'docente_creador': docente,
        },
        docente,
        stats,
    )
    if not creado:
        campos = []
        if first_name and not estudiante.first_name:
            estudiante.first_name = first_name
            campos.append('first_name')
        if last_name and not estudiante.last_name:
            estudiante.last_name = last_name
            campos.append('last_name')
        if campos:
            estudiante.save(update_fields=campos)

    if identificacion and estudiante.identificacion != identificacion:
        estudiante.identificacion = identificacion
        estudiante.save(update_fields=['identificacion'])

    materia = _resolver_materia(_celda_texto(fila.get('materia')), docente, stats)
    grupo_nombre = _celda_texto(fila.get('grupo'))
    if grupo_nombre:
        grupo, grupo_creado = Grupo.objects.get_or_create(
            docente=docente,
            nombre=grupo_nombre,
            defaults={'materia': materia, 'descripcion': ''},
        )
        if grupo_creado:
            stats['grupos_creados'] += 1
        else:
            campos = []
            if materia and grupo.materia_id != materia.id:
                grupo.materia = materia
                campos.append('materia')
            if campos:
                grupo.save(update_fields=campos)
        _inscribir_en_grupo(grupo, estudiante, stats)


def importar_estudiantes(archivo, docente) -> dict:
    if docente.rol not in (Usuario.Rol.DOCENTE, Usuario.Rol.ADMIN):
        raise ValidationError('Solo docentes pueden importar estudiantes.')

    filas = leer_filas_archivo(archivo)
    if not filas:
        raise ValidationError({'archivo': 'No hay filas de datos para importar.'})

    _sincronizar_secuencias_academico()

    stats = {
        'filas_procesadas': 0,
        'filas_exitosas': 0,
        'estudiantes_creados': 0,
        'estudiantes_vinculados': 0,
        'grupos_creados': 0,
        'inscripciones': 0,
        'materias_creadas': 0,
        'errores': [],
        'advertencias': [],
    }

    for fila in filas:
        stats['filas_procesadas'] += 1
        try:
            with transaction.atomic():
                _procesar_fila_estudiante(fila, docente, stats)
            stats['filas_exitosas'] += 1
        except ValidationError as exc:
            detalle = exc.detail
            if isinstance(detalle, dict):
                mensaje = '; '.join(
                    f'{k}: {", ".join(v if isinstance(v, list) else [str(v)])}'
                    for k, v in detalle.items()
                )
            elif isinstance(detalle, list):
                mensaje = '; '.join(str(x) for x in detalle)
            else:
                mensaje = str(detalle)
            stats['errores'].append({'fila': fila.get('_fila'), 'mensaje': mensaje})
        except Exception as exc:  # noqa: BLE001 — fila a fila, capturamos error genérico
            stats['errores'].append({'fila': fila.get('_fila'), 'mensaje': str(exc)})

    return stats


def _obtener_o_crear_por_correo(correo: str, docente, stats: dict) -> Estudiante:
    local = correo.split('@')[0].replace('.', ' ').replace('_', ' ').title()
    first_name, last_name = _partir_nombre(local)
    estudiante, _creado = _get_or_create_estudiante(
        correo,
        {
            'first_name': first_name or correo,
            'last_name': last_name,
            'docente_creador': docente,
        },
        docente,
        stats,
    )
    return estudiante


def _procesar_fila_grupo(fila: dict, docente, stats: dict) -> None:
    grupo_nombre = _celda_texto(fila.get('grupo'))
    if not grupo_nombre:
        raise ValidationError('La columna grupo es obligatoria.')

    materia = _resolver_materia(_celda_texto(fila.get('materia')), docente, stats)
    periodo = _celda_texto(fila.get('periodo'))

    grupo, creado = Grupo.objects.get_or_create(
        docente=docente,
        nombre=grupo_nombre,
        defaults={'materia': materia, 'periodo': periodo, 'descripcion': ''},
    )
    if creado:
        stats['grupos_creados'] += 1
    else:
        campos = []
        if materia and grupo.materia_id != materia.id:
            grupo.materia = materia
            campos.append('materia')
        if periodo and not grupo.periodo:
            grupo.periodo = periodo
            campos.append('periodo')
        if campos:
            grupo.save(update_fields=campos)

    raw_estudiantes = _celda_texto(fila.get('estudiantes'))
    if not raw_estudiantes:
        return

    for token in SEPARADOR_ESTUDIANTES.split(raw_estudiantes):
        token = token.strip()
        if not token:
            continue
        if '@' not in token:
            raise ValidationError(f'Estudiante inválido (usa correo): "{token}"')
        estudiante = _obtener_o_crear_por_correo(token.lower(), docente, stats)
        _inscribir_en_grupo(grupo, estudiante, stats)


def importar_grupos(archivo, docente) -> dict:
    if docente.rol not in (Usuario.Rol.DOCENTE, Usuario.Rol.ADMIN):
        raise ValidationError('Solo docentes pueden importar grupos.')

    filas = leer_filas_archivo(archivo)
    if not filas:
        raise ValidationError({'archivo': 'No hay filas de datos para importar.'})

    _sincronizar_secuencias_academico()

    stats = {
        'filas_procesadas': 0,
        'filas_exitosas': 0,
        'estudiantes_creados': 0,
        'estudiantes_vinculados': 0,
        'grupos_creados': 0,
        'inscripciones': 0,
        'materias_creadas': 0,
        'errores': [],
        'advertencias': [],
    }

    for fila in filas:
        stats['filas_procesadas'] += 1
        try:
            with transaction.atomic():
                _procesar_fila_grupo(fila, docente, stats)
            stats['filas_exitosas'] += 1
        except ValidationError as exc:
            detalle = exc.detail
            if isinstance(detalle, dict):
                mensaje = '; '.join(
                    f'{k}: {", ".join(v if isinstance(v, list) else [str(v)])}'
                    for k, v in detalle.items()
                )
            elif isinstance(detalle, list):
                mensaje = '; '.join(str(x) for x in detalle)
            else:
                mensaje = str(detalle)
            stats['errores'].append({'fila': fila.get('_fila'), 'mensaje': mensaje})
        except Exception as exc:  # noqa: BLE001
            stats['errores'].append({'fila': fila.get('_fila'), 'mensaje': str(exc)})

    return stats


def generar_plantilla_estudiantes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Estudiantes'
    ws.append(['nombre', 'correo', 'identificacion', 'materia', 'grupo'])
    ws.append([
        'Ana López',
        'ana.lopez@institucion.edu',
        '20201234',
        'Psicología social',
        'Grupo A',
    ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generar_plantilla_grupos() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Grupos'
    ws.append(['grupo', 'materia', 'periodo', 'estudiantes'])
    ws.append([
        'Grupo A',
        'Psicología social',
        '2026-1',
        'ana.lopez@institucion.edu; pedro.ramirez@institucion.edu',
    ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
