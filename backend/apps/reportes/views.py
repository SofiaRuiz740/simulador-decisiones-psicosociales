"""Endpoints de Reportes (admin métricas + descarga PDF/Excel)."""

import io
from datetime import datetime

from django.db.models import Avg, Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.academico.models import Estudiante, Grupo, Materia
from apps.casos.models import Caso
from apps.practicas.models import Practica
from apps.resultados.models import Resultado
from apps.usuarios.models import Usuario
from apps.usuarios.permissions import EsAdmin, EsDocenteOAdmin

from .serializers import DocenteAdminSerializer, EventoActividadSerializer
from .services import (
    REPORTES_TEMATICOS,
    analitica_reportes,
    datos_reporte_tematico,
    eventos_actividad,
    eventos_actividad_docente,
    filtrar_resultados,
    metricas_docente,
    resumen_reportes,
)


@api_view(['GET'])
@permission_classes([EsAdmin])
def admin_metricas(request):
    """Métricas globales para el dashboard del administrador (RF04, RF44)."""
    return Response({
        'docentes': Usuario.objects.filter(rol=Usuario.Rol.DOCENTE).count(),
        'estudiantes': Estudiante.objects.count(),
        'grupos': Grupo.objects.count(),
        'casos': Caso.objects.count(),
        'practicas': Practica.objects.count(),
        'practicas_por_estado': {
            estado: Practica.objects.filter(estado=estado).count()
            for estado, _ in Practica.Estado.choices
        },
        'resultados': Resultado.objects.count(),
        'nota_promedio': Resultado.objects.aggregate(avg=Avg('nota_final'))['avg'] or 0,
    })


@api_view(['GET'])
@permission_classes([EsAdmin])
def admin_docentes(request):
    """Listado de docentes con conteos agregados para el panel admin."""
    docentes = Usuario.objects.filter(rol=Usuario.Rol.DOCENTE).annotate(
        casos_count=Count('casos_creados', distinct=True),
        practicas_count=Count('practicas', distinct=True),
        estudiantes_count=Count('estudiantes', distinct=True),
        grupos_count=Count('grupos', distinct=True),
        materias_count=Count('materias', distinct=True),
    ).order_by('-date_joined')

    data = []
    for docente in docentes:
        data.append({
            'id': docente.id,
            'username': docente.username,
            'email': docente.email,
            'first_name': docente.first_name,
            'last_name': docente.last_name,
            'nombre_completo': docente.get_full_name() or docente.username,
            'is_active': docente.is_active,
            'date_joined': docente.date_joined,
            'casos_count': docente.casos_count,
            'practicas_count': docente.practicas_count,
            'estudiantes_count': docente.estudiantes_count,
            'grupos_count': docente.grupos_count,
            'materias_count': docente.materias_count,
        })
    return Response(DocenteAdminSerializer(data, many=True).data)


@api_view(['GET'])
@permission_classes([EsAdmin])
def admin_actividad(request):
    """Feed de actividad reciente del sistema (derivado de timestamps del dominio)."""
    try:
        limit = int(request.query_params.get('limit', 50))
    except (TypeError, ValueError):
        limit = 50
    limit = max(1, min(limit, 200))
    return Response(EventoActividadSerializer(eventos_actividad(limit=limit), many=True).data)


def _parse_filtros(request) -> dict:
    desde = request.query_params.get('desde') or None
    hasta = request.query_params.get('hasta') or None
    materia_id = request.query_params.get('materia_id') or None
    grupo_id = request.query_params.get('grupo_id') or None

    def _parse_date(val: str | None):
        if not val:
            return None
        try:
            return datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            return None

    def _parse_int(val: str | None):
        if not val:
            return None
        try:
            return int(val)
        except (TypeError, ValueError):
            return None

    return {
        'desde': _parse_date(desde),
        'hasta': _parse_date(hasta),
        'materia_id': _parse_int(materia_id),
        'grupo_id': _parse_int(grupo_id),
    }


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def docente_metricas(request):
    """Métricas del panel docente (casos, estudiantes, prácticas activas, feedback pendiente)."""
    return Response(metricas_docente(request.user))


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def docente_actividad(request):
    """Actividad reciente del docente autenticado."""
    try:
        limit = int(request.query_params.get('limit', 10))
    except (TypeError, ValueError):
        limit = 10
    limit = max(1, min(limit, 50))
    eventos = eventos_actividad_docente(request.user, limit=limit)
    return Response(EventoActividadSerializer(eventos, many=True).data)


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reportes_resumen(request):
    """Resumen agregado para la pantalla de reportes (métricas + filtros)."""
    filtros = _parse_filtros(request)
    return Response(resumen_reportes(request.user, **filtros))


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reportes_analitica(request):
    """Analítica agregada: desempeño por criterio, distribución de notas, feedback pendiente."""
    filtros = _parse_filtros(request)
    return Response(analitica_reportes(request.user, **filtros))


def _resultados_de_practica(practica_id: int, user: Usuario):
    qs = Resultado.objects.select_related(
        'participacion__estudiante', 'participacion__practica__caso',
    ).filter(participacion__practica_id=practica_id)
    if user.rol != Usuario.Rol.ADMIN:
        qs = qs.filter(participacion__practica__docente=user)
    return qs


def _assert_recurso_docente(user: Usuario, *, docente_id: int | None) -> Response | None:
    if user.rol == Usuario.Rol.DOCENTE and docente_id != user.id:
        return Response({'detail': 'No autorizado.'}, status=403)
    return None


def _respuesta_pdf(titulo: str, subtitulo: str, qs, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, title=titulo)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f'<b>{titulo}</b>', styles['Title']),
        Paragraph(subtitulo, styles['Normal']),
        Paragraph(f'Participantes: {qs.count()}', styles['Normal']),
        Spacer(1, 12),
    ]
    data = [['Estudiante', 'Correo', 'Práctica', 'Nota', 'Correctas', 'Incorrectas', 'Sin resp.']]
    for r in qs:
        data.append([
            r.participacion.estudiante.nombre_completo,
            r.participacion.estudiante.correo,
            r.participacion.practica.nombre,
            str(r.nota_final),
            r.correctas,
            r.incorrectas,
            r.no_respondidas,
        ])
    if len(data) == 1:
        elements.append(Paragraph('<i>Sin resultados aún.</i>', styles['Italic']))
    else:
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#283593')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ('ALIGN', (3, 1), (-1, -1), 'CENTER'),
        ]))
        elements.append(tbl)
    doc.build(elements)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _respuesta_excel(qs, filename: str) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resultados'
    ws.append(['Estudiante', 'Correo', 'Práctica', 'Nota', 'Correctas', 'Incorrectas', 'Sin resp.', 'Feedback'])
    for r in qs:
        ws.append([
            r.participacion.estudiante.nombre_completo,
            r.participacion.estudiante.correo,
            r.participacion.practica.nombre,
            float(r.nota_final),
            r.correctas, r.incorrectas, r.no_respondidas,
            r.feedback_docente,
        ])
    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_grupo_pdf(request, grupo_id: int):
    grupo = get_object_or_404(Grupo, pk=grupo_id)
    if err := _assert_recurso_docente(request.user, docente_id=grupo.docente_id):
        return err
    qs = filtrar_resultados(request.user, grupo_id=grupo_id)
    return _respuesta_pdf(
        f'Reporte por grupo: {grupo.nombre}',
        f'Docente: {grupo.docente.get_full_name() or grupo.docente.username}',
        qs,
        f'reporte-grupo-{grupo_id}.pdf',
    )


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_grupo_excel(request, grupo_id: int):
    grupo = get_object_or_404(Grupo, pk=grupo_id)
    if err := _assert_recurso_docente(request.user, docente_id=grupo.docente_id):
        return err
    qs = filtrar_resultados(request.user, grupo_id=grupo_id)
    return _respuesta_excel(qs, f'reporte-grupo-{grupo_id}.xlsx')


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_materia_pdf(request, materia_id: int):
    materia = get_object_or_404(Materia, pk=materia_id)
    if err := _assert_recurso_docente(request.user, docente_id=materia.docente_id):
        return err
    qs = filtrar_resultados(request.user, materia_id=materia_id)
    return _respuesta_pdf(
        f'Reporte por materia: {materia.nombre}',
        f'Docente: {materia.docente.get_full_name() or materia.docente.username}',
        qs,
        f'reporte-materia-{materia_id}.pdf',
    )


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_materia_excel(request, materia_id: int):
    materia = get_object_or_404(Materia, pk=materia_id)
    if err := _assert_recurso_docente(request.user, docente_id=materia.docente_id):
        return err
    qs = filtrar_resultados(request.user, materia_id=materia_id)
    return _respuesta_excel(qs, f'reporte-materia-{materia_id}.xlsx')


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_estudiante_pdf(request, estudiante_id: int):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
    if request.user.rol == Usuario.Rol.DOCENTE and not estudiante.docentes.filter(pk=request.user.pk).exists():
        return Response({'detail': 'No autorizado.'}, status=403)
    qs = filtrar_resultados(request.user, estudiante_id=estudiante_id)
    return _respuesta_pdf(
        f'Reporte por estudiante: {estudiante.nombre_completo}',
        f'Correo: {estudiante.correo}',
        qs,
        f'reporte-estudiante-{estudiante_id}.pdf',
    )


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_estudiante_excel(request, estudiante_id: int):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
    if request.user.rol == Usuario.Rol.DOCENTE and not estudiante.docentes.filter(pk=request.user.pk).exists():
        return Response({'detail': 'No autorizado.'}, status=403)
    qs = filtrar_resultados(request.user, estudiante_id=estudiante_id)
    return _respuesta_excel(qs, f'reporte-estudiante-{estudiante_id}.xlsx')


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_practica_pdf(request, practica_id: int):
    """GET /api/reportes/practica/{id}/pdf/ — descarga PDF con resultados de la práctica."""
    qs = _resultados_de_practica(practica_id, request.user)
    practica = get_object_or_404(Practica, pk=practica_id)
    if request.user.rol == Usuario.Rol.DOCENTE and practica.docente_id != request.user.id:
        return Response({'detail': 'No autorizado.'}, status=403)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, title=f'Reporte {practica.nombre}')
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f'<b>Reporte de práctica</b>: {practica.nombre}', styles['Title']),
        Paragraph(f'Caso: {practica.caso.nombre} · Docente: {practica.docente.get_full_name() or practica.docente.username}', styles['Normal']),
        Paragraph(f'Estado: {practica.get_estado_display()} · Participantes: {qs.count()}', styles['Normal']),
        Spacer(1, 12),
    ]
    data = [['Estudiante', 'Correo', 'Nota', 'Correctas', 'Incorrectas', 'Sin resp.']]
    for r in qs:
        data.append([
            r.participacion.estudiante.nombre_completo,
            r.participacion.estudiante.correo,
            str(r.nota_final),
            r.correctas,
            r.incorrectas,
            r.no_respondidas,
        ])
    if len(data) == 1:
        elements.append(Paragraph('<i>Sin resultados aún.</i>', styles['Italic']))
    else:
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#283593')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ]))
        elements.append(tbl)
    doc.build(elements)

    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="reporte-practica-{practica_id}.pdf"'
    return resp


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_practica_excel(request, practica_id: int):
    """GET /api/reportes/practica/{id}/excel/ — descarga XLSX con resultados."""
    qs = _resultados_de_practica(practica_id, request.user)
    practica = get_object_or_404(Practica, pk=practica_id)
    if request.user.rol == Usuario.Rol.DOCENTE and practica.docente_id != request.user.id:
        return Response({'detail': 'No autorizado.'}, status=403)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resultados'
    ws.append(['Estudiante', 'Correo', 'Nota', 'Correctas', 'Incorrectas', 'Sin resp.', 'Feedback'])
    for r in qs:
        ws.append([
            r.participacion.estudiante.nombre_completo,
            r.participacion.estudiante.correo,
            float(r.nota_final),
            r.correctas, r.incorrectas, r.no_respondidas,
            r.feedback_docente,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="reporte-practica-{practica_id}.xlsx"'
    return resp


# =============================================================================
# Reportes temáticos (P1) — Participación, Desempeño, Respuestas, Tiempos,
# Notas, Retroalimentaciones, Feedback pendiente.
# =============================================================================

def _renderizar_pdf(data: dict, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, title=data['titulo'])
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f'<b>{data["titulo"]}</b>', styles['Title']),
        Paragraph(data.get('subtitulo', ''), styles['Normal']),
        Spacer(1, 14),
    ]
    if not data['filas']:
        elements.append(Paragraph('<i>Aún no hay datos para este reporte.</i>', styles['Italic']))
    else:
        tabla_data = [data['columnas']] + [
            [str(v) if v is not None else '—' for v in fila] for fila in data['filas']
        ]
        tbl = Table(tabla_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(tbl)
    doc.build(elements)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _renderizar_excel(data: dict, filename: str) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = data['titulo'][:30]
    ws.append(data['columnas'])
    for fila in data['filas']:
        ws.append([v if v is not None else '' for v in fila])
    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _parse_filtros_tematico(request):
    def _date(name):
        raw = request.GET.get(name)
        if not raw:
            return None
        try:
            return datetime.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            return None

    def _int(name):
        raw = request.GET.get(name)
        try:
            return int(raw) if raw else None
        except (TypeError, ValueError):
            return None

    return {
        'desde': _date('desde'),
        'hasta': _date('hasta'),
        'materia_id': _int('materia_id'),
        'grupo_id': _int('grupo_id'),
        'estudiante_id': _int('estudiante_id'),
    }


@api_view(['GET'])
@permission_classes([EsDocenteOAdmin])
def reporte_tematico(request, tipo: str, formato: str):
    """GET /api/reportes/tematico/{tipo}/{formato}/

    `tipo`    ∈ participacion | desempeno | respuestas | tiempos | notas |
                retroalimentaciones | feedback
    `formato` ∈ pdf | excel
    Filtros opcionales: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD&materia_id=&grupo_id=&estudiante_id=
    """
    tipo = (tipo or '').lower()
    formato = (formato or '').lower()

    if tipo not in REPORTES_TEMATICOS:
        return Response(
            {'detail': f'Tipo de reporte inválido: "{tipo}".',
             'tipos_validos': list(REPORTES_TEMATICOS.keys())},
            status=400,
        )
    if formato not in {'pdf', 'excel'}:
        return Response(
            {'detail': 'Formato inválido. Debe ser "pdf" o "excel".'},
            status=400,
        )

    filtros = _parse_filtros_tematico(request)
    data = datos_reporte_tematico(tipo, request.user, **filtros)
    if data is None:
        return Response({'detail': 'Tipo no soportado.'}, status=400)

    ext = 'xlsx' if formato == 'excel' else 'pdf'
    filename = f'reporte-{tipo}.{ext}'
    return (_renderizar_excel if formato == 'excel' else _renderizar_pdf)(data, filename)
