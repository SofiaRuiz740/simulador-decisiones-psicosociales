"""Generación de plantillas estáticas para importación (caso, rúbrica, guía)."""

from __future__ import annotations

import io
import zipfile
from xml.sax.saxutils import escape

from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer


def _generar_docx(paragraphs: list[str]) -> bytes:
    """DOCX mínimo válido (sin dependencia python-docx)."""
    body = ''.join(
        f'<w:p><w:r><w:t xml:space="preserve">{escape(p)}</w:t></w:r></w:p>'
        for p in paragraphs
    )
    document_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>{body}<w:sectPr/></w:body>
</w:document>'''
    content_types = '''<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml"
    ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>'''
    rels = '''<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
    Target="word/document.xml"/>
</Relationships>'''

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', content_types)
        zf.writestr('_rels/.rels', rels)
        zf.writestr('word/document.xml', document_xml)
    return buffer.getvalue()


def generar_plantilla_caso() -> bytes:
    return _generar_docx([
        'PLANTILLA DE CASO — Simulador de decisiones psicosociales',
        '',
        '=== DATOS GENERALES ===',
        'Nombre: [Nombre del caso]',
        'Área psicosocial: [Ej. Convivencia escolar]',
        'Tiempo estimado (min): [30]',
        '',
        '=== CONTEXTO E HISTORIA ===',
        'Describa la situación, actores y antecedentes del caso…',
        '',
        '=== ESCENARIO 1 ===',
        'Título: [Título del escenario]',
        'Narrativa: [Descripción de la situación que enfrentará el estudiante]',
        '',
        'Pregunta 1: [Enunciado de la pregunta]',
        'A) [Opción A]',
        'B) [Opción B]',
        'C) [Opción C]',
        'Respuesta correcta: [A/B/C]',
        '',
        '=== RÚBRICA (referencia) ===',
        'Criterio 1 — Empatía (peso 30%)',
        'Criterio 2 — Análisis (peso 40%)',
        'Criterio 3 — Propuesta (peso 30%)',
    ])


def generar_caso_ejemplo() -> bytes:
    return _generar_docx([
        'CASO EJEMPLO — Conflicto en el aula',
        '',
        '=== DATOS GENERALES ===',
        'Nombre: Conflicto entre compañeros en clase de psicología',
        'Área psicosocial: Convivencia escolar',
        'Tiempo estimado (min): 25',
        '',
        '=== CONTEXTO E HISTORIA ===',
        'En un curso de psicología social, dos estudiantes han tenido un altercado '
        'durante una exposición. El docente solicita apoyo para mediar la situación '
        'sin escalar el conflicto.',
        '',
        '=== ESCENARIO 1 ===',
        'Título: Primer acercamiento',
        'Narrativa: Llegas al salón y observas tensión entre Ana y Pedro. '
        'Varios estudiantes miran expectantes.',
        '',
        'Pregunta 1: ¿Cuál es la primera acción más adecuada?',
        'A) Separar inmediatamente a ambos sin escuchar',
        'B) Invitar a cada uno a expresar su versión en un espacio calmado',
        'C) Ignorar el conflicto para continuar la clase',
        'Respuesta correcta: B',
    ])


def generar_plantilla_rubrica() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rubrica'
    ws.append([
        'criterio', 'descripcion', 'peso',
        'nivel_1_nombre', 'nivel_1_descriptor',
        'nivel_2_nombre', 'nivel_2_descriptor',
        'nivel_3_nombre', 'nivel_3_descriptor',
        'nivel_4_nombre', 'nivel_4_descriptor',
    ])
    ws.append([
        'Empatía',
        'Capacidad de identificar emociones del otro',
        30,
        'Incipiente', 'No reconoce emociones ajenas',
        'En desarrollo', 'Reconoce emociones básicas',
        'Logrado', 'Identifica emociones y contexto',
        'Sobresaliente', 'Demuestra comprensión profunda',
    ])
    ws.append([
        'Análisis',
        'Capacidad de analizar la situación psicosocial',
        40,
        'Incipiente', 'Análisis superficial',
        'En desarrollo', 'Identifica algunos factores',
        'Logrado', 'Analiza factores relevantes',
        'Sobresaliente', 'Análisis integral y fundamentado',
    ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parsear_plantilla_rubrica(archivo_excel) -> dict:
    """Lee un XLSX con el formato de `generar_plantilla_rubrica()` y devuelve
    `{'criterios': [...], 'errores': [...], 'advertencias': [...]}` listo
    para asignarse a `Rubrica.criterios`.

    Columnas requeridas: criterio, descripcion, peso,
                         nivel_1_nombre, nivel_1_descriptor,
                         nivel_2_nombre, nivel_2_descriptor,
                         nivel_3_nombre, nivel_3_descriptor,
                         nivel_4_nombre, nivel_4_descriptor.

    Acepta hasta 4 niveles (extiende si la plantilla añade nivel_5 en futuro).
    """
    from openpyxl import load_workbook

    errores: list[dict] = []
    advertencias: list[dict] = []

    if hasattr(archivo_excel, 'seek'):
        archivo_excel.seek(0)

    try:
        wb = load_workbook(archivo_excel, data_only=True)
    except Exception as exc:
        return {
            'criterios': [],
            'errores': [{'fila': None, 'mensaje': f'No se pudo abrir el Excel: {exc}'}],
            'advertencias': [],
        }

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {
            'criterios': [],
            'errores': [{'fila': None, 'mensaje': 'El archivo está vacío.'}],
            'advertencias': [],
        }

    encabezados = [str(c or '').strip().lower() for c in rows[0]]
    requeridas = {'criterio', 'peso'}
    faltantes = requeridas - set(encabezados)
    if faltantes:
        return {
            'criterios': [],
            'errores': [{
                'fila': 1,
                'mensaje': f'Faltan columnas obligatorias: {", ".join(sorted(faltantes))}.',
            }],
            'advertencias': [],
        }

    def col(nombre: str) -> int | None:
        try:
            return encabezados.index(nombre)
        except ValueError:
            return None

    criterios = []
    for nro_fila, fila in enumerate(rows[1:], start=2):
        nombre = (fila[col('criterio')] or '').strip() if col('criterio') is not None else ''
        if not nombre:
            continue  # fila vacía, se ignora

        desc_idx = col('descripcion')
        peso_idx = col('peso')
        descripcion = (fila[desc_idx] or '').strip() if desc_idx is not None and desc_idx < len(fila) else ''

        try:
            peso = int(fila[peso_idx] or 0) if peso_idx is not None and peso_idx < len(fila) else 0
        except (TypeError, ValueError):
            errores.append({'fila': nro_fila, 'mensaje': f'Peso inválido en criterio "{nombre}".'})
            peso = 0

        niveles = []
        for n in range(1, 5):
            cn = col(f'nivel_{n}_nombre')
            cd = col(f'nivel_{n}_descriptor')
            nombre_n = (fila[cn] or '').strip() if cn is not None and cn < len(fila) else ''
            descriptor_n = (fila[cd] or '').strip() if cd is not None and cd < len(fila) else ''
            if nombre_n or descriptor_n:
                niveles.append({
                    'nivel': n,
                    'nombre': nombre_n or f'Nivel {n}',
                    'descriptor': descriptor_n,
                })

        if not niveles:
            advertencias.append({
                'fila': nro_fila,
                'mensaje': f'Criterio "{nombre}" no tiene niveles; se aplicarán los globales.',
            })

        criterios.append({
            'id': f'c{nro_fila - 1}',
            'nombre': nombre,
            'descripcion': descripcion,
            'peso': peso,
            'niveles': niveles,
        })

    suma_pesos = sum(c['peso'] for c in criterios)
    if criterios and suma_pesos != 100:
        advertencias.append({
            'fila': None,
            'mensaje': (
                f'La suma de pesos es {suma_pesos}, no 100. '
                'La rúbrica se guardará pero el cálculo de nota usará el modo plano.'
            ),
        })

    return {'criterios': criterios, 'errores': errores, 'advertencias': advertencias}


def generar_guia_importacion() -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title='Guía de importación')
    styles = getSampleStyleSheet()
    elements = [
        Paragraph('<b>Guía de importación</b>', styles['Title']),
        Paragraph('Simulador de decisiones psicosociales', styles['Normal']),
        Spacer(1, 12),
        Paragraph('<b>1. Estudiantes (Excel/CSV)</b>', styles['Heading2']),
        Paragraph(
            'Columnas: nombre, correo, identificacion, materia, grupo. '
            'Descarga la plantilla desde la pestaña Plantillas.',
            styles['Normal'],
        ),
        Spacer(1, 8),
        Paragraph('<b>2. Grupos (Excel/CSV)</b>', styles['Heading2']),
        Paragraph(
            'Columnas: grupo, materia, periodo, estudiantes (correos separados por ;).',
            styles['Normal'],
        ),
        Spacer(1, 8),
        Paragraph('<b>3. Casos (DOCX/PDF/TXT)</b>', styles['Heading2']),
        Paragraph(
            'Sube un documento con datos generales, contexto, escenarios y preguntas. '
            'Usa la plantilla de caso como referencia de estructura.',
            styles['Normal'],
        ),
        Spacer(1, 8),
        Paragraph('<b>4. Rúbrica (Excel)</b>', styles['Heading2']),
        Paragraph(
            'La plantilla de rúbrica define criterios, pesos y niveles de desempeño. '
            'Puedes editarla y usarla como referencia al configurar la rúbrica del caso.',
            styles['Normal'],
        ),
    ]
    doc.build(elements)
    return buffer.getvalue()
